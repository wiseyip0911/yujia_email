#!/usr/bin/env python3
import argparse
import csv
import hashlib
import json
import sys
import time
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from backend.sci_platform.config import DATA_DIR
from backend.sci_platform.db import get_connection, initialize_database
from backend.sci_platform.mailbox_testing import (
    ConnectionResult,
    mask_email,
    normalize_email,
    provider_for_email,
    test_imap_login,
)


@dataclass(frozen=True)
class MailboxRow:
    source_row: int
    project_code: str
    customer_name: str
    author_name: str
    email: str
    password: str


def cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_rows(path: Path) -> list[MailboxRow]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl is required. Use the bundled Python runtime shown in README.md.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [cell_text(value) for value in rows[0]]
    required = ["项目编号", "客户名称", "作者名字", "邮箱"]
    index = {}
    for name in required:
        if name in header:
            index[name] = header.index(name)
    missing_headers = sorted(set(required) - set(index))
    if missing_headers:
        raise SystemExit(f"Missing required headers: {', '.join(missing_headers)}")
    credential_headers = ["授权码", "授权密码", "邮箱授权码", "客户端授权码", "密码"]
    credential_indexes = [header.index(name) for name in credential_headers if name in header]
    if not credential_indexes:
        raise SystemExit("Missing credential header: expected one of 授权码/授权密码/邮箱授权码/客户端授权码/密码")

    parsed = []
    for row_number, row in enumerate(rows[1:], start=2):
        values = list(row)
        if not any(cell_text(value) for value in values):
            continue
        parsed.append(
            MailboxRow(
                source_row=row_number,
                project_code=cell_text(values[index["项目编号"]]),
                customer_name=cell_text(values[index["客户名称"]]),
                author_name=cell_text(values[index["作者名字"]]),
                email=normalize_email(cell_text(values[index["邮箱"]])),
                password=first_filled_cell(values, credential_indexes),
            )
        )
    return parsed


def first_filled_cell(values: list[object], indexes: list[int]) -> str:
    for index in indexes:
        if index < len(values):
            value = cell_text(values[index])
            if value:
                return value
    return ""


def password_fingerprint(password: str) -> str:
    return hashlib.sha256(password.encode("utf-8")).hexdigest()


def import_metadata(conn, rows: list[MailboxRow], source_file: str, import_batch: str) -> dict:
    imported_rows = 0
    skipped_missing_credentials = 0
    mailbox_ids = {}

    for row in rows:
        if not row.email or not row.password:
            skipped_missing_credentials += 1
            continue

        customer_name = row.customer_name or "待补客户资料"
        customer = conn.execute("SELECT * FROM customers WHERE name = ?", (customer_name,)).fetchone()
        if not customer:
            conn.execute(
                """
                INSERT INTO customers (name, contact_name, owner, status)
                VALUES (?, ?, ?, ?)
                """,
                (customer_name, row.author_name or customer_name, "运营一组", "active"),
            )
            customer_id = conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
        else:
            customer_id = customer["customer_id"]

        provider = provider_for_email(row.email)
        existing = conn.execute("SELECT mailbox_id FROM mailboxes WHERE email_address = ?", (row.email,)).fetchone()
        if existing:
            mailbox_id = existing["mailbox_id"]
            conn.execute(
                """
                UPDATE mailboxes
                SET mailbox_type = ?, auth_method = ?, status = CASE WHEN status = 'active' THEN status ELSE 'pending_test' END
                WHERE mailbox_id = ?
                """,
                ("IMAP", provider.auth_method, mailbox_id),
            )
        else:
            conn.execute(
                """
                INSERT INTO mailboxes (customer_id, email_address, mailbox_type, auth_method, status)
                VALUES (?, ?, ?, ?, ?)
                """,
                (customer_id, row.email, "IMAP", provider.auth_method, "pending_test"),
            )
            mailbox_id = conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]

        mailbox_ids[row.email] = mailbox_id
        conn.execute(
            """
            INSERT INTO mailbox_project_links
                (mailbox_id, project_code, customer_name, author_name, source_file, source_row, import_batch)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(source_file, source_row) DO UPDATE SET
                mailbox_id = excluded.mailbox_id,
                project_code = excluded.project_code,
                customer_name = excluded.customer_name,
                author_name = excluded.author_name,
                import_batch = excluded.import_batch
            """,
            (
                mailbox_id,
                row.project_code,
                row.customer_name,
                row.author_name,
                source_file,
                row.source_row,
                import_batch,
            ),
        )
        imported_rows += 1

    return {
        "imported_rows": imported_rows,
        "skipped_missing_credentials": skipped_missing_credentials,
        "unique_mailboxes": len(mailbox_ids),
        "mailbox_ids": mailbox_ids,
    }


def group_credentials(rows: list[MailboxRow]) -> dict[str, list[MailboxRow]]:
    grouped = defaultdict(list)
    seen = set()
    for row in rows:
        if not row.email or not row.password:
            continue
        key = (row.email, password_fingerprint(row.password))
        if key in seen:
            continue
        seen.add(key)
        grouped[row.email].append(row)
    return dict(grouped)


def status_from_result(result: ConnectionResult) -> tuple[str, Optional[str]]:
    if result.result == "success":
        return "active", None
    if result.error_type == "needs_microsoft_oauth2":
        return "needs_oauth", "Microsoft mailbox requires OAuth2 connection"
    if result.error_type == "needs_google_oauth_or_app_password":
        return "needs_oauth", "Gmail requires Sign in with Google or an app password"
    if result.error_type == "needs_manual_imap_config":
        return "config_required", "Manual IMAP host/port is required"
    if result.error_type == "auth_failed":
        return "auth_failed", result.error_message
    if result.error_type == "security_blocked":
        return "security_blocked", result.error_message
    return "test_failed", result.error_message


def save_test_result(
    conn,
    batch: str,
    mailbox_id: int,
    provider_name: str,
    imap_host: Optional[str],
    imap_port: Optional[int],
    variant: str,
    result: ConnectionResult,
) -> None:
    conn.execute(
        """
        INSERT INTO mailbox_connection_tests
            (test_batch, mailbox_id, provider, imap_host, imap_port, credential_variant, result, error_type, error_message, inbox_message_count)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            batch,
            mailbox_id,
            provider_name,
            imap_host,
            imap_port,
            variant,
            result.result,
            result.error_type,
            result.error_message,
            result.inbox_message_count,
        ),
    )
    status, reason = status_from_result(result)
    if result.result == "success":
        conn.execute(
            """
            UPDATE mailboxes
            SET status = ?, error_reason = NULL, last_sync_at = CURRENT_TIMESTAMP
            WHERE mailbox_id = ?
            """,
            (status, mailbox_id),
        )
    else:
        preserve_active = result.error_type in {"network_timeout", "network_error", "dns_error"}
        conn.execute(
            """
            UPDATE mailboxes
            SET status = CASE WHEN status = 'active' AND ? THEN status ELSE ? END,
                error_reason = ?
            WHERE mailbox_id = ?
            """,
            (1 if preserve_active else 0, status, reason, mailbox_id),
        )


def run_tests(
    conn,
    batch: str,
    grouped: dict[str, list[MailboxRow]],
    mailbox_ids: dict[str, int],
    timeout: int,
    delay: float,
    limit: Optional[int],
) -> list[dict]:
    reports = []
    attempts = 0
    for email in sorted(grouped):
        rows = grouped[email]
        config = provider_for_email(email)
        for variant_index, row in enumerate(rows, start=1):
            if limit is not None and attempts >= limit:
                return reports
            attempts += 1
            variant = f"v{variant_index}"
            result = test_imap_login(email, row.password, config, timeout_seconds=timeout)
            mailbox_id = mailbox_ids[email]
            save_test_result(conn, batch, mailbox_id, config.provider, config.imap_host, config.imap_port, variant, result)
            reports.append(
                {
                    "email": email,
                    "masked_email": mask_email(email),
                    "provider": config.provider,
                    "imap_host": config.imap_host or "",
                    "credential_variant": variant,
                    "result": result.result,
                    "error_type": result.error_type or "",
                    "error_message": result.error_message or "",
                    "inbox_message_count": result.inbox_message_count,
                    "source_rows": str(row.source_row),
                }
            )
            print(
                f"[{attempts}] {mask_email(email)} {config.provider} {variant}: "
                f"{result.result}{' / ' + result.error_type if result.error_type else ''}",
                flush=True,
            )
            if delay > 0:
                time.sleep(delay)
            if result.result == "success":
                break
    return reports


def write_reports(reports: list[dict], import_summary: dict, output_dir: Path, batch: str) -> dict:
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = output_dir / f"mailbox-connection-{batch}.csv"
    json_path = output_dir / f"mailbox-connection-{batch}.json"

    fields = [
        "masked_email",
        "provider",
        "imap_host",
        "credential_variant",
        "result",
        "error_type",
        "error_message",
        "inbox_message_count",
        "source_rows",
    ]
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in reports:
            writer.writerow({field: row.get(field, "") for field in fields})

    result_counts = defaultdict(int)
    error_counts = defaultdict(int)
    provider_counts = defaultdict(int)
    for report in reports:
        result_counts[report["result"]] += 1
        if report["error_type"]:
            error_counts[report["error_type"]] += 1
        provider_counts[report["provider"]] += 1

    payload = {
        "batch": batch,
        "import_summary": {k: v for k, v in import_summary.items() if k != "mailbox_ids"},
        "attempt_count": len(reports),
        "result_counts": dict(sorted(result_counts.items())),
        "error_counts": dict(sorted(error_counts.items())),
        "provider_counts": dict(sorted(provider_counts.items())),
        "csv_report": str(csv_path),
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    payload["json_report"] = str(json_path)
    return payload


def main() -> int:
    parser = argparse.ArgumentParser(description="Import mailbox metadata and run conservative IMAP connection tests.")
    parser.add_argument("--xlsx", required=True, help="Source workbook with 项目编号/客户名称/作者名字/邮箱/密码 columns.")
    parser.add_argument("--timeout", type=int, default=15, help="Per-mailbox IMAP timeout seconds.")
    parser.add_argument("--delay", type=float, default=0.8, help="Delay between login attempts.")
    parser.add_argument("--limit", type=int, default=None, help="Optional max credential variants to test.")
    parser.add_argument("--dry-run", action="store_true", help="Import metadata and summarize without external login attempts.")
    args = parser.parse_args()

    source_path = Path(args.xlsx).expanduser().resolve()
    batch = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    rows = load_rows(source_path)
    initialize_database()

    with get_connection() as conn:
        import_summary = import_metadata(conn, rows, source_path.name, batch)
        grouped = group_credentials(rows)
        if args.dry_run:
            provider_counts = defaultdict(int)
            for email in grouped:
                provider_counts[provider_for_email(email).provider] += len(grouped[email])
            print(
                json.dumps(
                    {
                        "batch": batch,
                        "import_summary": {k: v for k, v in import_summary.items() if k != "mailbox_ids"},
                        "credential_variants": sum(len(items) for items in grouped.values()),
                        "provider_counts": dict(sorted(provider_counts.items())),
                    },
                    ensure_ascii=False,
                    indent=2,
                )
            )
            return 0

        reports = run_tests(
            conn,
            batch,
            grouped,
            import_summary["mailbox_ids"],
            timeout=args.timeout,
            delay=args.delay,
            limit=args.limit,
        )
        summary = write_reports(reports, import_summary, DATA_DIR / "connection_tests", batch)
        print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
